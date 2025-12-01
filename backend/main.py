import os
import re
import io
import pathlib
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from pymongo import ReturnDocument
from .database import db
from .models import UserCreate, TokenResponse, VerifyEmailRequest, CategoryCreate, RaporCreate, UserAdminCreate, UserListItem, ProjectCreate
from .auth import create_access_token, verify_password, hash_password, get_current_user, require_role

app = FastAPI(title="EKOS")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

upload_root = pathlib.Path(__file__).parent / "uploads"
upload_root.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(upload_root)), name="uploads")

async def ensure_admin():
    admin_email = "ibrahimznrmak@gmail.com"
    admin_username = "miharbirnz"
    admin_password = "279eS490"
    exists = await db.users.find_one({"email": admin_email})
    if not exists:
        await db.users.insert_one({
            "username": admin_username,
            "email": admin_email,
            "hashed_password": hash_password(admin_password),
            "role": "admin",
            "is_active": True,
            "is_email_verified": True,
        })
    else:
        await db.users.update_one({"_id": exists["_id"]}, {"$set": {
            "username": admin_username,
            "hashed_password": hash_password(admin_password),
            "role": "admin",
            "is_active": True,
            "is_email_verified": True,
        }})

@app.on_event("startup")
async def startup_event():
    await ensure_admin()

def oid(s: str) -> ObjectId:
    try:
        return ObjectId(s)
    except Exception:
        raise HTTPException(status_code=400, detail="Geçersiz id")

async def await_counter(prefix: str) -> int:
    doc = await db.counters.find_one_and_update(
        {"_id": prefix}, {"$inc": {"seq": 1}}, upsert=True, return_document=ReturnDocument.AFTER
    )
    return int(doc.get("seq", 1))

async def generate_rapor_no(city_code: str) -> str:
    year = datetime.utcnow().year
    prefix = f"PK{year}-{city_code}"
    counter = await await_counter(prefix)
    return f"{prefix}{counter:03d}"

@app.post("/auth/register")
async def register(payload: UserCreate):
    if payload.password != payload.password_confirm:
        raise HTTPException(status_code=400, detail="Şifreler uyuşmuyor")
    existing = await db.users.find_one({"$or": [{"username": payload.username}, {"email": payload.email}]})
    if existing:
        raise HTTPException(status_code=400, detail="Kullanıcı mevcut")
    token = os.urandom(16).hex()
    await db.users.insert_one({
        "username": payload.username,
        "email": payload.email,
        "hashed_password": hash_password(payload.password),
        "role": "inspector",
        "is_active": True,
        "is_email_verified": False,
        "email_token": token,
    })
    return {"message": "Kayıt başarılı", "verification_token": token}

@app.post("/auth/verify-email")
async def verify_email(payload: VerifyEmailRequest):
    user = await db.users.find_one({"email": payload.email})
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    if user.get("email_token") != payload.token:
        raise HTTPException(status_code=400, detail="Token geçersiz")
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"is_email_verified": True}, "$unset": {"email_token": ""}})
    return {"message": "E-posta doğrulandı"}

@app.post("/auth/login", response_model=TokenResponse)
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    identifier = form_data.username
    user = await db.users.find_one({"$or": [{"username": identifier}, {"email": identifier}]})
    if not user:
        raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
    if not verify_password(form_data.password, user.get("hashed_password", "")):
        raise HTTPException(status_code=401, detail="Şifre hatalı")
    if not user.get("is_email_verified"):
        raise HTTPException(status_code=403, detail="E-posta doğrulanmamış")
    token = create_access_token({"sub": user["username"], "role": user.get("role", "viewer")})
    return {"access_token": token, "token_type": "bearer"}

@app.get("/raporlar")
async def list_raporlar(
    q: Optional[str] = None,
    kategori_id: Optional[str] = None,
    periyot: Optional[int] = None,
    uygunluk: Optional[str] = None,
    project_id: Optional[str] = None,
    user = Depends(get_current_user),
):
    conds = []
    if q:
        conds.append({
            "$or": [
                {"rapor_no": {"$regex": q, "$options": "i"}},
                {"ekipman_adi": {"$regex": q, "$options": "i"}},
                {"firma": {"$regex": q, "$options": "i"}},
            ]
        })
    if kategori_id:
        conds.append({"kategori_id": kategori_id})
    if project_id:
        conds.append({"project_id": project_id})
    if periyot:
        conds.append({"periyot": periyot})
    if uygunluk:
        if uygunluk == "uygun":
            conds.append({"uygunluk_durumu": {"$regex": "^uygun$", "$options": "i"}})
        elif uygunluk == "uygun_degil":
            conds.append({
                "$or": [
                    {"uygunluk_durumu": {"$regex": "^uygun_degil$", "$options": "i"}},
                    {"uygunluk_durumu": {"$regex": "^uygun değil$", "$options": "i"}},
                    {"uygunluk_durumu": {"$regex": "^uygun degil$", "$options": "i"}},
                    {"uygunluk_durumu": {"$regex": "^uygunsuz$", "$options": "i"}},
                ]
            })
    filt = {"$and": conds} if conds else {}
    cur = db.raporlar.find(filt).sort("_id", -1)
    items = []
    async for r in cur:
        r["id"] = str(r.pop("_id"))
        items.append(r)
    return {"items": items}

@app.post("/raporlar")
async def create_rapor(payload: RaporCreate, user = Depends(require_role("inspector"))):
    if not payload.ekipman_adi or not payload.kategori_id or not payload.firma:
        raise HTTPException(status_code=400, detail="Zorunlu alanlar eksik")
    city_src = payload.project_city or payload.lokasyon or "IST"
    city_code = str(city_src).upper()[:3]
    rapor_no = await generate_rapor_no(city_code)
    doc = payload.model_dump()
    if payload.project_id:
        try:
            pr = await db.projeler.find_one({"_id": oid(payload.project_id)})
            if pr:
                doc["project_name"] = pr.get("name")
        except Exception:
            pass
    doc.update({
        "rapor_no": rapor_no,
        "olusturan_username": user["username"],
        "created_at": datetime.utcnow().isoformat(),
    })
    res = await db.raporlar.insert_one(doc)
    return {"id": str(res.inserted_id), "rapor_no": rapor_no}

@app.get("/raporlar/{id}")
async def get_rapor(id: str, user = Depends(get_current_user)):
    r = await db.raporlar.find_one({"_id": oid(id)})
    if not r:
        raise HTTPException(status_code=404, detail="Bulunamadı")
    r["id"] = str(r.pop("_id"))
    return r

@app.put("/raporlar/{id}")
async def update_rapor(id: str, payload: RaporCreate, user = Depends(require_role("inspector"))):
    await db.raporlar.update_one({"_id": oid(id)}, {"$set": payload.model_dump()})
    return {"message": "Güncellendi"}

@app.delete("/raporlar/{id}")
async def delete_rapor(id: str, user = Depends(require_role("inspector"))):
    await db.raporlar.delete_one({"_id": oid(id)})
    return {"message": "Silindi"}

@app.get("/users")
async def list_users(user = Depends(require_role("admin"))):
    items = []
    async for u in db.users.find({}).sort("username", 1):
        items.append({
            "id": str(u.get("_id")),
            "username": u.get("username"),
            "email": u.get("email"),
            "role": u.get("role", "inspector"),
        })
    return {"items": items}

@app.post("/users")
async def create_user(payload: UserAdminCreate, user = Depends(require_role("admin"))):
    existing = await db.users.find_one({"$or": [{"username": payload.username}, {"email": payload.email}]})
    if existing:
        raise HTTPException(status_code=400, detail="Kullanıcı mevcut")
    await db.users.insert_one({
        "username": payload.username,
        "email": payload.email,
        "hashed_password": hash_password(payload.password),
        "role": payload.role or "inspector",
        "is_active": True,
        "is_email_verified": True,
    })
    return {"message": "Oluşturuldu"}

@app.get("/projeler")
async def list_projects(user = Depends(get_current_user)):
    items = []
    async for p in db.projeler.find({}).sort("name", 1):
        p["id"] = str(p.pop("_id"))
        items.append(p)
    return {"items": items}

@app.post("/projeler")
async def create_project(payload: ProjectCreate, user = Depends(require_role("admin"))):
    res = await db.projeler.insert_one({"name": payload.name})
    return {"id": str(res.inserted_id)}

@app.put("/projeler/{id}")
async def update_project(id: str, payload: ProjectCreate, user = Depends(require_role("admin"))):
    await db.projeler.update_one({"_id": oid(id)}, {"$set": {"name": payload.name}})
    return {"message": "Güncellendi"}

@app.delete("/projeler/{id}")
async def delete_project(id: str, user = Depends(require_role("admin"))):
    await db.projeler.delete_one({"_id": oid(id)})
    return {"message": "Silindi"}

@app.get("/kategoriler")
async def list_categories(user = Depends(get_current_user)):
    items = []
    async for c in db.kategoriler.find({}).sort("name", 1):
        c["id"] = str(c.pop("_id"))
        items.append(c)
    unique = {}
    for c in items:
        unique[c["name"].strip().lower()] = c
    return {"items": list(unique.values())}

@app.post("/kategoriler")
async def create_category(payload: CategoryCreate, user = Depends(require_role("admin"))):
    res = await db.kategoriler.insert_one(payload.model_dump())
    return {"id": str(res.inserted_id)}

@app.put("/kategoriler/{id}")
async def update_category(id: str, payload: CategoryCreate, user = Depends(require_role("admin"))):
    await db.kategoriler.update_one({"_id": oid(id)}, {"$set": payload.model_dump()})
    return {"message": "Güncellendi"}

@app.delete("/kategoriler/{id}")
async def delete_category(id: str, user = Depends(require_role("admin"))):
    await db.kategoriler.delete_one({"_id": oid(id)})
    return {"message": "Silindi"}

@app.put("/users/{id}")
async def update_user(id: str, payload: dict = Body(...), user = Depends(require_role("admin"))):
    updates = {}
    for key in ["username", "email", "role", "is_active", "is_email_verified"]:
        if key in payload:
            updates[key] = payload[key]
    if "password" in payload and payload["password"]:
        updates["hashed_password"] = hash_password(payload["password"]) 
    if not updates:
        raise HTTPException(status_code=400, detail="Güncellenecek alan yok")
    await db.users.update_one({"_id": oid(id)}, {"$set": updates})
    return {"message": "Güncellendi"}

@app.delete("/users/{id}")
async def delete_user(id: str, user = Depends(require_role("admin"))):
    await db.users.delete_one({"_id": oid(id)})
    return {"message": "Silindi"}

@app.get("/me")
async def get_me(user = Depends(get_current_user)):
    return {
        "id": str(user.get("_id")),
        "username": user.get("username"),
        "email": user.get("email"),
        "role": user.get("role", "inspector"),
        "full_name": user.get("full_name"),
        "avatar_url": user.get("avatar_url"),
    }

@app.put("/me")
async def update_me(payload: dict = Body(...), user = Depends(get_current_user)):
    updates = {}
    for key in ["email", "full_name"]:
        if key in payload:
            updates[key] = payload[key]
    if "password" in payload and payload["password"]:
        updates["hashed_password"] = hash_password(payload["password"]) 
    if not updates:
        raise HTTPException(status_code=400, detail="Güncellenecek alan yok")
    await db.users.update_one({"_id": user["_id"]}, {"$set": updates})
    return {"message": "Güncellendi"}

@app.post("/me/avatar")
async def upload_avatar(file: UploadFile = File(...), user = Depends(get_current_user)):
    if file.content_type not in ["image/jpeg", "image/png"]:
        raise HTTPException(status_code=400, detail="Desteklenmeyen format")
    size_hdr = file.headers.get("content-length")
    if size_hdr and int(size_hdr) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya limiti aşıldı")
    udir = upload_root / "users" / str(user["_id"]) 
    udir.mkdir(parents=True, exist_ok=True)
    target = udir / file.filename
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya limiti aşıldı")
    target.write_bytes(data)
    url = f"/uploads/users/{str(user['_id'])}/{file.filename}"
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"avatar_url": url}})
    return {"avatar_url": url}

@app.post("/upload")
async def upload_file(
    rapor_id: str = Form(...),
    file: UploadFile = File(...),
    user = Depends(get_current_user),
):
    if file.content_type not in ["image/jpeg", "image/png", "application/pdf"]:
        raise HTTPException(status_code=400, detail="Desteklenmeyen format")
    size_hdr = file.headers.get("content-length")
    if size_hdr and int(size_hdr) > 4 * 1024 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya limiti aşıldı")
    rid = oid(rapor_id)
    rdir = upload_root / str(rid)
    rdir.mkdir(parents=True, exist_ok=True)
    target = rdir / file.filename
    data = await file.read()
    if len(data) > 4 * 1024 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya limiti aşıldı")
    target.write_bytes(data)
    media = {
        "filename": file.filename,
        "content_type": file.content_type,
        "size": len(data),
        "url": f"/uploads/{rid}/{file.filename}",
    }
    await db.raporlar.update_one({"_id": rid}, {"$push": {"medyalar": media}})
    return {"message": "Yüklendi", "media": media}

@app.get("/export-excel")
async def export_excel(template: Optional[bool] = None, user = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    headers = [
        "Rapor No","Ekipman Adı","Kategori","Firma","Lokasyon","Marka/Model","Seri No","Alt Kategori","Periyot","Geçerlilik Tarihi","Açıklama","Uygunluk","Proje","Proje İli","Proje İlçesi","Oluşturma/Yüklenme Tarihi"
    ]
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i).font = Font(bold=True)
    if not template:
        async for r in db.raporlar.find({}).sort("_id", -1):
            created_dt = None
            try:
                created_dt = r.get("imported_at") or r.get("created_at") or r.get("_id").generation_time.isoformat()
            except Exception:
                created_dt = None
            ws.append([
                r.get("rapor_no"), r.get("ekipman_adi"), r.get("kategori_id"), r.get("firma"), r.get("lokasyon"), r.get("marka_model"), r.get("seri_no"), r.get("alt_kategori"), r.get("periyot"), r.get("gecerlilik_tarihi"), r.get("aciklama"), r.get("uygunluk_durumu"), r.get("project_name") or r.get("project_id"), r.get("project_city"), r.get("project_district"), created_dt
            ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "ekos-template.xlsx" if template else "ekos-raporlar.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    project_id: Optional[str] = Form(None),
    user = Depends(require_role("inspector"))
):
    if file.content_type not in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"]:
        raise HTTPException(status_code=400, detail="Excel dosyası gerekli")
    data = await file.read()
    stream = io.BytesIO(data)
    from openpyxl import load_workbook
    wb = load_workbook(stream)
    ws = wb.active
    imported = 0
    form_pid = None
    form_pname = None
    if project_id:
        try:
            pid = oid(project_id)
            pr = await db.projeler.find_one({"_id": pid})
            if pr:
                form_pid = str(pid)
                form_pname = pr.get("name")
        except Exception:
            pr = await db.projeler.find_one({"name": project_id})
            if pr:
                form_pid = str(pr.get("_id"))
                form_pname = pr.get("name")
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            continue
        try:
            ekipman_adi = row[1]
            kategori_id = str(row[2]) if row[2] else ""
            firma = row[3]
            if not ekipman_adi or not kategori_id or not firma:
                continue
            raw_uyg = (str(row[11]).strip().lower() if row[11] else "")
            if raw_uyg:
                if any(x in raw_uyg for x in ["degil", "değil", "uygunsuz"]):
                    raw_uyg = "uygun_degil"
                else:
                    raw_uyg = "uygun"
            raw_periyot = row[8]
            periyot_val = None
            if raw_periyot is not None:
                try:
                    periyot_val = int(raw_periyot)
                except Exception:
                    m = re.search(r"\d+", str(raw_periyot))
                    periyot_val = int(m.group(0)) if m else None
            payload = {
                "ekipman_adi": ekipman_adi,
                "kategori_id": kategori_id,
                "firma": firma,
                "lokasyon": row[4],
                "marka_model": row[5],
                "seri_no": row[6],
                "alt_kategori": row[7],
                "periyot": periyot_val,
                "gecerlilik_tarihi": row[9],
                "aciklama": row[10],
                "uygunluk_durumu": raw_uyg,
                "olusturan_username": user["username"],
            }
            project_cell = row[12] if len(row) > 12 else None
            if form_pid:
                payload["project_id"] = form_pid
                if form_pname:
                    payload["project_name"] = form_pname
            elif project_cell:
                val = str(project_cell).strip()
                try:
                    pid = oid(val)
                    payload["project_id"] = str(pid)
                    pr = await db.projeler.find_one({"_id": pid})
                    if pr:
                        payload["project_name"] = pr.get("name")
                except Exception:
                    pr = await db.projeler.find_one({"name": val})
                    if pr:
                        payload["project_id"] = str(pr.get("_id"))
                        payload["project_name"] = pr.get("name")
            payload["project_city"] = row[13] if len(row) > 13 else None
            payload["project_district"] = row[14] if len(row) > 14 else None
            city_src = payload.get("project_city") or payload.get("lokasyon") or "IST"
            city_code = str(city_src).upper()[:3]
            rapor_no = await generate_rapor_no(city_code)
            payload.update({"rapor_no": rapor_no, "imported_at": datetime.utcnow().isoformat()})
            await db.raporlar.insert_one(payload)
            imported += 1
        except Exception:
            continue
    return {"imported": imported}

@app.get("/")
async def root():
    return {"name": "EKOS"}
